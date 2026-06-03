"""Bank accounts, statements, and cash-flow endpoints."""
import csv
import io
import os
import re
import secrets
from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional

import pdfplumber
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.middleware.auth import get_current_user
from app.middleware.rbac import get_effective_tenant_id
from app.models.user import User

router = APIRouter(prefix="/bank", tags=["bank"])


# ── Pydantic models ───────────────────────────────────────────────────────────

class AccountCreate(BaseModel):
    name: str
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    opening_balance: float = 0.0
    currency: str = "MYR"


class AccountOut(BaseModel):
    id: int
    name: str
    bank_name: Optional[str]
    account_number: Optional[str]
    opening_balance: float
    currency: str
    current_balance: float
    total_credit: float
    total_debit: float


class CategoryCreate(BaseModel):
    name: str
    type: str = "expense"  # income | expense
    color: str = "#6366f1"


class CategoryOut(BaseModel):
    id: int
    name: str
    type: str
    color: str


class CategoryRef(BaseModel):
    id: int
    name: str
    color: str


class InvoiceRef(BaseModel):
    id: int
    invoice_number: str
    client_name: Optional[str] = None


class TransactionCreate(BaseModel):
    txn_date: date
    description: str
    party_name: Optional[str] = None
    amount: float
    type: str  # credit | debit
    category_ids: List[int] = []
    invoice_ids: List[int] = []
    note: Optional[str] = None


class TransactionUpdate(BaseModel):
    txn_date: Optional[date] = None
    description: Optional[str] = None
    party_name: Optional[str] = None
    amount: Optional[float] = None
    type: Optional[str] = None
    category_ids: Optional[List[int]] = None  # None = unchanged; [] = clear all
    invoice_ids: Optional[List[int]] = None   # None = unchanged; [] = clear all
    invoice_id: Optional[int] = None          # backward-compat (ignored if invoice_ids set)
    bill_id: Optional[int] = None
    note: Optional[str] = None


class TransactionOut(BaseModel):
    id: int
    account_id: int
    statement_id: Optional[int]
    txn_date: str
    description: str
    party_name: Optional[str]
    amount: float
    type: str
    categories: List[CategoryRef] = []
    # backward-compat single-category fields (first category or None)
    category_id: Optional[int] = None
    category_name: Optional[str] = None
    category_color: Optional[str] = None
    invoices: List[InvoiceRef] = []
    # backward-compat single-invoice fields (first invoice or None)
    invoice_id: Optional[int] = None
    invoice_number: Optional[str] = None
    bill_id: Optional[int]
    bill_number: Optional[str]
    note: Optional[str]
    receipt_url: Optional[str] = None


class ParsedRow(BaseModel):
    txn_date: str
    description: str
    party_name: Optional[str] = None
    amount: float
    type: str  # credit | debit
    note: Optional[str] = None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _tenant_id(user: User) -> Optional[int]:
    return user.tenant_id


async def _get_account(db: AsyncSession, account_id: int, tenant_id: Optional[int]):
    r = await db.execute(
        text("SELECT id, tenant_id FROM bank_accounts WHERE id = :id"),
        {"id": account_id},
    )
    row = r.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Account not found")
    if tenant_id and row[1] != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")
    return row


async def _account_balances(db: AsyncSession, account_id: int, opening: float):
    r = await db.execute(
        text("""
            SELECT
                COALESCE(SUM(CASE WHEN type='credit' THEN amount ELSE 0 END), 0) AS total_credit,
                COALESCE(SUM(CASE WHEN type='debit'  THEN amount ELSE 0 END), 0) AS total_debit
            FROM bank_transactions WHERE account_id = :aid
        """),
        {"aid": account_id},
    )
    row = r.fetchone()
    total_credit = float(row[0])
    total_debit = float(row[1])
    current_balance = opening + total_credit - total_debit
    return total_credit, total_debit, current_balance


_MAX_AMOUNT = Decimal("9999999999999.99")  # DECIMAL(15,2) upper bound


def _parse_amount(raw: str) -> Optional[float]:
    if not raw:
        return None
    raw = str(raw).strip()
    negative = raw.startswith("-") or (raw.startswith("(") and raw.endswith(")"))
    cleaned = re.sub(r"[^\d.]", "", raw)
    if not cleaned:
        return None
    # Multiple dots means ambiguous format — keep only last decimal group
    parts = cleaned.split(".")
    if len(parts) > 2:
        cleaned = "".join(parts[:-1]) + "." + parts[-1]
    try:
        val = float(cleaned)
        # Reject amounts that overflow DECIMAL(15,2)
        if val > float(_MAX_AMOUNT):
            return None
        return -val if negative else val
    except (ValueError, OverflowError):
        return None


def _parse_date(raw: str) -> Optional[date]:
    raw = str(raw).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d %b %Y", "%d %B %Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    # DDMMYYYY — CIMB Portfolio/CASA format (e.g. 26052026)
    if re.fullmatch(r"\d{8}", raw):
        try:
            return datetime.strptime(raw, "%d%m%Y").date()
        except ValueError:
            pass
    return None


def _detect_column(headers: List[str], candidates: List[str]) -> Optional[int]:
    """Match column headers using whole-word logic to avoid false positives.

    e.g. "cr" must not match "description" (des-cr-iption).
    """
    for idx, h in enumerate(headers):
        # Split header into individual words, stripping punctuation
        h_words = set(re.split(r"[^a-z0-9]+", h.lower()))
        h_words.discard("")
        for c in candidates:
            c_lower = c.lower()
            # Exact word match
            if c_lower in h_words:
                return idx
            # Prefix match: "deposit" matches "deposits", "withdrawal" matches "withdrawals"
            if any(w.startswith(c_lower) for w in h_words):
                return idx
            # Long-substring fallback (≥5 chars) — safe since short abbrevs handled above
            if len(c_lower) >= 5 and c_lower in h.lower():
                return idx
    return None


def _normalise_header(h: str) -> str:
    """Flatten multi-line pdfplumber cell values to single-line."""
    return re.sub(r"\s+", " ", str(h or "")).strip()


def _detect_type_col(headers: List[str]) -> Optional[int]:
    """Find a column whose header indicates credit/debit TYPE (not an amount)."""
    for idx, h in enumerate(headers):
        hn = _normalise_header(h).lower()
        if any(k in hn for k in ["amount type", "cr/dr", "dr/cr", "transaction type", "type"]):
            return idx
    return None


def _resolve_type(type_cell: str, debit_val: Optional[float], credit_val: Optional[float]) -> Optional[str]:
    """Determine 'credit' or 'debit' from a C/D type cell or separate debit/credit values."""
    if type_cell:
        t = type_cell.strip().upper()
        if t in ("C", "CR", "CREDIT"):
            return "credit"
        if t in ("D", "DR", "DEBIT"):
            return "debit"
    if credit_val and credit_val > 0:
        return "credit"
    if debit_val and debit_val > 0:
        return "debit"
    return None


def _parse_cimb_portfolio_text(raw: str) -> List[ParsedRow]:
    """
    Parse CIMB Portfolio / CASA account statement from raw PDF text.

    Each row follows: ACCT(10) SEQ(9) DATE(8,DDMMYYYY) CODE(3-4) DESCRIPTION BRANCH DOCREF
    MYR AMOUNT C/D MYR BALANCE BALTYPE TIME CUSTREF 1 RECIPREF OTHERREF SENDERNAME
    """
    # Fix hyphenated line-breaks ("MDN2605260-\n14801826" → "MDN2605260-14801826")
    text = re.sub(r"(\w)-\s+(\w)", r"\1-\2", raw)
    text = re.sub(r"\s+", " ", text)

    main_re = re.compile(
        r"\d{10}\s+"            # account number
        r"\d{9}\s+"             # record sequence
        r"(\d{8})\s+"           # date DDMMYYYY  [1]
        r"\d{3,4}\s+"           # transaction code
        r"([A-Z][A-Z ]+?)\s+"   # description (uppercase) [2]
        r"(?:-|\d{4})\s+"       # originating branch code
        r"(\S+)\s+"             # document reference [3]
        r"MYR\s*([\d,. ]+?)\s+" # transaction amount [4]
        r"([CD])\s+"            # amount type  [5]
        r"MYR"                  # balance start (anchor, not captured)
    )

    matches = list(main_re.finditer(text))
    results: List[ParsedRow] = []

    for i, m in enumerate(matches):
        parsed_date = _parse_date(m.group(1))
        if not parsed_date:
            continue

        amount = _parse_amount(re.sub(r"\s+", "", m.group(4)))
        if not amount:
            continue

        txn_type = "credit" if m.group(5) == "C" else "debit"
        description = m.group(2).strip()
        doc_ref = m.group(3)

        # ── Extract customer reference and sender name from the tail ──────────
        tail_end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        tail = text[m.end(): tail_end]

        # Customer reference: between "D time(6)" and filler "1 "
        cust_ref: Optional[str] = None
        cr_match = re.search(r"[CD]\s+\d{6}\s+(.*?)\s+1\s+", tail)
        if cr_match:
            cr_raw = cr_match.group(1).strip()
            if cr_raw and cr_raw != "-":
                cust_ref = cr_raw

        # Sender name: after filler "1 " — skip up to two reference tokens, then grab name
        sender: Optional[str] = None
        sn_match = re.search(r"\s1\s+\S+\s+(?:\S+\s+)?([A-Z][A-Z0-9\s\.\'/\-]+)", tail)
        if sn_match:
            raw_sn = re.sub(r"\d{8,}.*$", "", sn_match.group(1)).strip()
            if raw_sn:
                sender = raw_sn

        # note: prefer descriptive customer ref over raw doc ref
        if cust_ref and not re.fullmatch(r"[\dA-Za-z/\-]+", cust_ref):
            note: Optional[str] = cust_ref
        elif doc_ref and doc_ref != "-":
            note = doc_ref
        else:
            note = cust_ref  # may be None

        results.append(ParsedRow(
            txn_date=str(parsed_date),
            description=description,
            party_name=sender or None,
            amount=round(abs(amount), 2),
            type=txn_type,
            note=note,
        ))

    return results


def _parse_cimb_portfolio_page(page) -> List[ParsedRow]:
    """
    Parse a single CIMB Portfolio/CASA statement page.

    Each logical table row spans multiple physical lines (cells wrap: "DUITNOW TO" /
    "ACCOUNT", "MYR 16,000." / "00", etc.).  The fix: find every occurrence of the
    10-digit account number in the leftmost column — each marks exactly one transaction.
    Collect ALL words between consecutive account-number y-positions, sort by (y, x),
    then apply the transaction regex.  This reconstructs rows correctly regardless of
    column-major text ordering or wrapped cells.
    """
    try:
        words = page.extract_words(x_tolerance=5, y_tolerance=3)
    except Exception:
        return _parse_cimb_portfolio_text(page.extract_text() or "")

    if not words:
        return _parse_cimb_portfolio_text(page.extract_text() or "")

    # ── 1. Find account-number words (10-digit integers) in the leftmost column ──
    acct_words = [w for w in words if re.fullmatch(r"\d{10}", w["text"])]
    if not acct_words:
        return _parse_cimb_portfolio_text(page.extract_text() or "")

    # Keep only the leftmost column (real acct-no column), skip any that appear
    # further right (shouldn't happen, but guards against edge cases).
    min_x0 = min(w["x0"] for w in acct_words)
    row_anchors = sorted(
        [w for w in acct_words if w["x0"] <= min_x0 + 15],
        key=lambda w: w["top"],
    )

    if not row_anchors:
        return _parse_cimb_portfolio_text(page.extract_text() or "")

    # ── 2. For each transaction (anchor → next anchor), collect & sort words ────
    row_re = re.compile(
        r"\d{10}\s+"            # account number
        r"\d{9}\s+"             # record sequence
        r"(\d{8})\s+"           # date DDMMYYYY  [1]
        r"\d{3,4}\s+"           # transaction code
        r"([A-Z][A-Z ]+?)\s+"   # description (uppercase, lazy) [2]
        r"(?:-|\d{4})\s+"       # originating branch code
        r"(\S+?)\s+"            # document reference (lazy) [3]
        r"MYR\s*([\d,. ]+?)\s+" # transaction amount (allows split like "16,000. ") [4]
        r"([CD])\s+"            # amount type  [5]
        r"MYR"                  # balance start anchor (not captured)
    )

    results: List[ParsedRow] = []

    for i, anchor in enumerate(row_anchors):
        y_start = anchor["top"] - 3
        y_end   = row_anchors[i + 1]["top"] - 3 if i + 1 < len(row_anchors) else float("inf")

        row_words = [w for w in words if y_start <= w["top"] < y_end]
        # Sort by (y-band, x) so the primary "first line" of each column comes first,
        # followed by wrapped second-line content.
        row_words = sorted(row_words, key=lambda w: (round(w["top"]), w["x0"]))
        row_text  = " ".join(w["text"] for w in row_words)

        m = row_re.search(row_text)
        if not m:
            continue

        parsed_date = _parse_date(m.group(1))
        if not parsed_date:
            continue

        amount = _parse_amount(re.sub(r"\s+", "", m.group(4)))
        if not amount:
            continue

        txn_type    = "credit" if m.group(5) == "C" else "debit"
        description = m.group(2).strip()
        doc_ref     = m.group(3).rstrip("-")   # strip trailing hyphen from split refs
        tail        = row_text[m.end():]

        # Customer reference: after time(6 digits) up to filler "1 "
        cust_ref: Optional[str] = None
        cr_match = re.search(r"[CD]\s+\d{6}\s+(.*?)\s+1\s+", tail)
        if cr_match:
            cr_raw = cr_match.group(1).strip()
            if cr_raw and cr_raw != "-":
                cust_ref = cr_raw

        # Sender name: last block of UPPER-CASE words in the tail (after filler "1 ")
        sender: Optional[str] = None
        # Look for the filler "1" then skip two reference tokens, then grab the name
        sn_match = re.search(
            r"\b1\b\s+\S.*?([A-Z]{2}[A-Z0-9\s\.\'/\-]+?)(?:\s+\d{6,}|\s*$)", tail
        )
        if sn_match:
            raw_sn = sn_match.group(1).strip()
            # Filter out short or purely numeric fragments
            if raw_sn and not re.fullmatch(r"[\d\-/]+", raw_sn) and len(raw_sn) >= 3:
                sender = raw_sn

        if cust_ref and not re.fullmatch(r"[\dA-Za-z/\-]+", cust_ref):
            note: Optional[str] = cust_ref
        elif doc_ref and doc_ref != "-":
            note = doc_ref
        else:
            note = cust_ref

        results.append(ParsedRow(
            txn_date=str(parsed_date),
            description=description,
            party_name=sender or None,
            amount=round(abs(amount), 2),
            type=txn_type,
            note=note,
        ))

    return results


def _detect_delimiter(content: str) -> str:
    """Pick the delimiter that produces the most columns in the first data row."""
    first_line = content.split("\n")[0]
    best = ","
    best_count = 0
    for delim in [",", ";", "\t", "|"]:
        row = next(csv.reader([first_line], delimiter=delim), [])
        if len(row) > best_count:
            best_count = len(row)
            best = delim
    return best
    return None


def _parse_csv_content(content: str) -> List[ParsedRow]:
    delimiter = _detect_delimiter(content)
    reader = csv.reader(io.StringIO(content), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return []

    # Find header row (first row with recognisable column names)
    header_idx = 0
    headers: List[str] = []
    for i, row in enumerate(rows[:15]):
        joined = " ".join(row).lower()
        if any(k in joined for k in ["date", "debit", "credit", "amount", "narration", "description", "transaction"]):
            header_idx = i
            headers = [_normalise_header(h) for h in row]
            break

    if not headers:
        return []

    date_col   = _detect_column(headers, ["date", "transaction date", "value date"])
    desc_col   = _detect_column(headers, ["description", "narration", "particular", "detail", "remark",
                                          "transaction code description", "code description"])
    debit_col  = _detect_column(headers, ["debit", "withdrawal", "dr"])
    credit_col = _detect_column(headers, ["credit", "deposit", "cr"])
    amount_col = _detect_column(headers, ["transaction amount", "amount"]) if (debit_col is None or credit_col is None) else None
    type_col   = _detect_type_col(headers)
    party_col  = _detect_column(headers, ["sender name", "sender", "party", "beneficiary", "payee", "remitter"])
    doc_col    = _detect_column(headers, ["document reference", "doc ref", "document ref"])
    cust_ref_col = _detect_column(headers, ["customer reference", "cust ref", "customer ref"])

    if date_col is None:
        return []

    def _cell(row: List[str], idx: Optional[int]) -> str:
        return row[idx].strip() if idx is not None and idx < len(row) else ""

    results: List[ParsedRow] = []
    for row in rows[header_idx + 1:]:
        if not row or all(c.strip() == "" for c in row):
            continue
        try:
            raw_date = _cell(row, date_col)
            parsed_date = _parse_date(raw_date)
            if not parsed_date:
                continue

            description = _cell(row, desc_col)
            party_name  = _cell(row, party_col) or None
            doc_ref     = _cell(row, doc_col)
            cust_ref    = _cell(row, cust_ref_col)
            note_parts  = [p for p in [doc_ref, cust_ref] if p]
            note        = " | ".join(note_parts) or None

            if type_col is not None and amount_col is not None:
                # CIMB-style: single amount column + C/D type column
                type_cell = _cell(row, type_col)
                amt = _parse_amount(_cell(row, amount_col))
                if amt is None or amt == 0:
                    continue
                txn_type = _resolve_type(type_cell, None, None)
                if txn_type is None:
                    continue
                amount = abs(amt)
            elif amount_col is not None:
                amt = _parse_amount(_cell(row, amount_col))
                if amt is None:
                    continue
                txn_type = "credit" if amt >= 0 else "debit"
                amount = abs(amt)
            else:
                debit  = _parse_amount(_cell(row, debit_col))  if debit_col  is not None else None
                credit = _parse_amount(_cell(row, credit_col)) if credit_col is not None else None
                txn_type = _resolve_type("", debit, credit)
                if txn_type is None:
                    continue
                amount = abs(debit) if txn_type == "debit" else abs(credit)  # type: ignore[arg-type]

            results.append(ParsedRow(
                txn_date=str(parsed_date),
                description=description,
                party_name=party_name,
                amount=round(amount, 2),
                type=txn_type,
                note=note,
            ))
        except (IndexError, Exception):
            continue

    return results


def _parse_pdf_content(file_bytes: bytes) -> List[ParsedRow]:
    results: List[ParsedRow] = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                page_had_table = False

                for table in tables:
                    if not table:
                        continue

                    # Find header row (search first 8 rows to handle multi-row headers)
                    header_row_idx = None
                    headers: List[str] = []
                    for i, row in enumerate(table[:8]):
                        if row is None:
                            continue
                        cells = [_normalise_header(c) for c in row]
                        joined = " ".join(cells).lower()
                        if any(k in joined for k in ["date", "debit", "credit", "amount", "narration",
                                                      "description", "transaction"]):
                            header_row_idx = i
                            headers = cells
                            break

                    if header_row_idx is None or not headers:
                        continue

                    date_col     = _detect_column(headers, ["transaction date", "date", "value date"])
                    desc_col     = _detect_column(headers, ["transaction code description", "code description",
                                                            "description", "narration", "particular", "detail"])
                    debit_col    = _detect_column(headers, ["debit", "withdrawal", "dr"])
                    credit_col   = _detect_column(headers, ["credit", "deposit", "cr"])
                    amount_col   = _detect_column(headers, ["transaction amount", "amount"]) \
                                   if (debit_col is None or credit_col is None) else None
                    type_col     = _detect_type_col(headers)
                    party_col    = _detect_column(headers, ["sender name", "sender", "beneficiary", "payee", "remitter"])
                    doc_col      = _detect_column(headers, ["document reference", "doc reference", "doc ref"])
                    cust_ref_col = _detect_column(headers, ["customer reference", "cust reference", "cust ref"])

                    if date_col is None:
                        continue

                    def _cell(row: List, idx: Optional[int]) -> str:
                        return _normalise_header(row[idx]) if idx is not None and idx < len(row) else ""

                    for row in table[header_row_idx + 1:]:
                        if not row or all((c is None or str(c).strip() == "") for c in row):
                            continue
                        try:
                            raw_date = _cell(row, date_col)
                            parsed_date = _parse_date(raw_date)
                            if not parsed_date:
                                continue

                            description  = _cell(row, desc_col)
                            party_name   = _cell(row, party_col) or None
                            doc_ref      = _cell(row, doc_col)
                            cust_ref     = _cell(row, cust_ref_col)
                            note_parts   = [p for p in [doc_ref, cust_ref] if p]
                            note         = " | ".join(note_parts) or None

                            if type_col is not None and amount_col is not None:
                                # CIMB Portfolio/CASA: single amount + C/D type column
                                type_cell = _cell(row, type_col)
                                amt = _parse_amount(_cell(row, amount_col))
                                if amt is None or amt == 0:
                                    continue
                                txn_type = _resolve_type(type_cell, None, None)
                                if txn_type is None:
                                    continue
                                amount = abs(amt)
                            elif amount_col is not None:
                                amt = _parse_amount(_cell(row, amount_col))
                                if amt is None:
                                    continue
                                txn_type = "credit" if amt >= 0 else "debit"
                                amount = abs(amt)
                            else:
                                debit  = _parse_amount(_cell(row, debit_col))  if debit_col  is not None else None
                                credit = _parse_amount(_cell(row, credit_col)) if credit_col is not None else None
                                txn_type = _resolve_type("", debit, credit)
                                if txn_type is None:
                                    continue
                                amount = abs(debit) if txn_type == "debit" else abs(credit)  # type: ignore[arg-type]

                            page_had_table = True
                            results.append(ParsedRow(
                                txn_date=str(parsed_date),
                                description=description,
                                party_name=party_name,
                                amount=round(amount, 2),
                                type=txn_type,
                                note=note,
                            ))
                        except Exception:
                            continue

                # Fallback: text-based parsing for statements without detectable table borders
                if not page_had_table:
                    text_content = page.extract_text() or ""

                    # CIMB Portfolio/CASA detection: statement has 10-digit account numbers and "Amount Type" column.
                    # Use the word-coordinate parser so rows are reconstructed left-to-right, not column-major.
                    if re.search(r"\d{10}", text_content) and re.search(r"Amount\s*Type|MYR\s+[\d,]+\.\d{2}\s+[CD]", text_content):
                        cimb_rows = _parse_cimb_portfolio_page(page)
                        if cimb_rows:
                            results.extend(cimb_rows)
                            continue  # skip generic regex for this page

                    # Generic regex for DD/MM/YYYY or DD-MM-YYYY statements
                    pattern = re.compile(
                        r"(\d{2}[\/\-]\d{2}[\/\-]\d{4})\s+([\w\s\-\/\*]+?)\s+([\d,]+\.\d{2})(?:\s+([\d,]+\.\d{2}))?",
                        re.MULTILINE,
                    )
                    for m in pattern.finditer(text_content):
                        parsed_date = _parse_date(m.group(1))
                        if not parsed_date:
                            continue
                        description = m.group(2).strip()
                        amt1 = _parse_amount(m.group(3))
                        amt2 = _parse_amount(m.group(4)) if m.group(4) else None
                        if amt2 is not None:
                            if amt1 and amt1 > 0:
                                results.append(ParsedRow(txn_date=str(parsed_date), description=description, amount=round(amt1, 2), type="debit"))
                            if amt2 and amt2 > 0:
                                results.append(ParsedRow(txn_date=str(parsed_date), description=description, amount=round(amt2, 2), type="credit"))
                        elif amt1:
                            results.append(ParsedRow(txn_date=str(parsed_date), description=description, amount=round(amt1, 2), type="debit"))
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Could not parse PDF: {e}")

    return results


# ── Category endpoints ────────────────────────────────────────────────────────

@router.get("/categories", response_model=List[CategoryOut])
async def list_categories(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    r = await db.execute(
        text("SELECT id, name, type, color FROM transaction_categories WHERE tenant_id = :tid ORDER BY type, name"),
        {"tid": tid},
    )
    return [CategoryOut(id=row[0], name=row[1], type=row[2], color=row[3]) for row in r.fetchall()]


@router.post("/categories", response_model=CategoryOut)
async def create_category(
    payload: CategoryCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    r = await db.execute(
        text("INSERT INTO transaction_categories (tenant_id, name, type, color) VALUES (:tid, :name, :type, :color)"),
        {"tid": tid, "name": payload.name, "type": payload.type, "color": payload.color},
    )
    await db.commit()
    cat_id = r.lastrowid
    return CategoryOut(id=cat_id, name=payload.name, type=payload.type, color=payload.color)


@router.put("/categories/{cat_id}", response_model=CategoryOut)
async def update_category(
    cat_id: int,
    payload: CategoryCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    check = await db.execute(
        text("SELECT id FROM transaction_categories WHERE id = :id AND tenant_id = :tid"),
        {"id": cat_id, "tid": tid},
    )
    if not check.fetchone():
        raise HTTPException(status_code=404, detail="Category not found")
    await db.execute(
        text("UPDATE transaction_categories SET name=:name, type=:type, color=:color WHERE id=:id"),
        {"name": payload.name, "type": payload.type, "color": payload.color, "id": cat_id},
    )
    await db.commit()
    return CategoryOut(id=cat_id, name=payload.name, type=payload.type, color=payload.color)


@router.delete("/categories/{cat_id}")
async def delete_category(
    cat_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    check = await db.execute(
        text("SELECT id FROM transaction_categories WHERE id = :id AND tenant_id = :tid"),
        {"id": cat_id, "tid": tid},
    )
    if not check.fetchone():
        raise HTTPException(status_code=404, detail="Category not found")
    await db.execute(text("DELETE FROM transaction_categories WHERE id = :id"), {"id": cat_id})
    await db.commit()
    return {"ok": True}


# ── Bank account endpoints ────────────────────────────────────────────────────

@router.get("/accounts", response_model=List[AccountOut])
async def list_accounts(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    r = await db.execute(
        text("SELECT id, name, bank_name, account_number, opening_balance, currency FROM bank_accounts WHERE tenant_id = :tid ORDER BY id"),
        {"tid": tid},
    )
    out = []
    for row in r.fetchall():
        acc_id, name, bank_name, acc_no, opening, currency = row
        tc, td, cb = await _account_balances(db, acc_id, float(opening))
        out.append(AccountOut(
            id=acc_id, name=name, bank_name=bank_name, account_number=acc_no,
            opening_balance=float(opening), currency=currency,
            current_balance=round(cb, 2), total_credit=round(tc, 2), total_debit=round(td, 2),
        ))
    return out


@router.post("/accounts", response_model=AccountOut)
async def create_account(
    payload: AccountCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    r = await db.execute(
        text("""
            INSERT INTO bank_accounts (tenant_id, name, bank_name, account_number, opening_balance, currency)
            VALUES (:tid, :name, :bank_name, :acc_no, :opening, :currency)
        """),
        {"tid": tid, "name": payload.name, "bank_name": payload.bank_name,
         "acc_no": payload.account_number, "opening": payload.opening_balance, "currency": payload.currency},
    )
    await db.commit()
    acc_id = r.lastrowid
    return AccountOut(
        id=acc_id, name=payload.name, bank_name=payload.bank_name,
        account_number=payload.account_number, opening_balance=payload.opening_balance,
        currency=payload.currency, current_balance=payload.opening_balance,
        total_credit=0.0, total_debit=0.0,
    )


@router.put("/accounts/{account_id}", response_model=AccountOut)
async def update_account(
    account_id: int,
    payload: AccountCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    await _get_account(db, account_id, tid)
    await db.execute(
        text("""
            UPDATE bank_accounts SET name=:name, bank_name=:bank_name,
            account_number=:acc_no, opening_balance=:opening, currency=:currency
            WHERE id=:id
        """),
        {"name": payload.name, "bank_name": payload.bank_name, "acc_no": payload.account_number,
         "opening": payload.opening_balance, "currency": payload.currency, "id": account_id},
    )
    await db.commit()
    tc, td, cb = await _account_balances(db, account_id, payload.opening_balance)
    return AccountOut(
        id=account_id, name=payload.name, bank_name=payload.bank_name,
        account_number=payload.account_number, opening_balance=payload.opening_balance,
        currency=payload.currency, current_balance=round(cb, 2),
        total_credit=round(tc, 2), total_debit=round(td, 2),
    )


@router.delete("/accounts/{account_id}")
async def delete_account(
    account_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    await _get_account(db, account_id, tid)
    await db.execute(text("DELETE FROM bank_accounts WHERE id = :id"), {"id": account_id})
    await db.commit()
    return {"ok": True}


# ── Statement upload & parse ──────────────────────────────────────────────────

@router.post("/debug/parse-pdf")
async def debug_parse_pdf(file: UploadFile = File(...)):
    """Temporary debug endpoint — returns raw pdfplumber extraction info."""
    import pdfplumber, io as _io
    content = await file.read()
    result: dict = {"pages": []}
    with pdfplumber.open(_io.BytesIO(content)) as pdf:
        for pg_num, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            words = page.extract_words(x_tolerance=5, y_tolerance=3)
            acct_words = [w for w in words if __import__("re").fullmatch(r"\d{10}", w["text"])]
            cimb_rows = _parse_cimb_portfolio_page(page)
            result["pages"].append({
                "page": pg_num + 1,
                "text_preview": text[:500],
                "word_count": len(words),
                "acct_number_hits": [{"text": w["text"], "x0": w["x0"], "top": w["top"]} for w in acct_words],
                "cimb_rows_found": len(cimb_rows),
                "cimb_rows": [r.dict() for r in cimb_rows],
            })
    return result


@router.post("/accounts/{account_id}/upload/preview")
async def preview_statement(
    account_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Parse a PDF or CSV and return extracted rows for user review — nothing saved yet."""
    tid = _tenant_id(current_user)
    await _get_account(db, account_id, tid)

    content = await file.read()
    filename = file.filename or ""

    if filename.lower().endswith(".pdf"):
        rows = _parse_pdf_content(content)
    elif filename.lower().endswith((".csv", ".txt")):
        rows = _parse_csv_content(content.decode("utf-8", errors="replace"))
    else:
        raise HTTPException(status_code=400, detail="Only PDF or CSV files are supported")

    if not rows:
        raise HTTPException(status_code=422, detail="No transactions could be extracted from this file. Try CSV export instead.")

    return {"rows": [r.dict() for r in rows], "count": len(rows)}


@router.post("/accounts/{account_id}/upload/confirm")
async def confirm_statement(
    account_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Save uploaded file + all extracted transactions to the database."""
    tid = _tenant_id(current_user)
    await _get_account(db, account_id, tid)

    content = await file.read()
    filename = file.filename or "statement"

    if filename.lower().endswith(".pdf"):
        rows = _parse_pdf_content(content)
    elif filename.lower().endswith((".csv", ".txt")):
        rows = _parse_csv_content(content.decode("utf-8", errors="replace"))
    else:
        raise HTTPException(status_code=400, detail="Only PDF or CSV files are supported")

    # Save file
    upload_dir = os.environ.get("UPLOAD_DIR", "uploads")
    safe_name = f"{account_id}_{secrets.token_hex(6)}_{filename}"
    file_path = f"{upload_dir}/bank_statements/{safe_name}"
    with open(file_path, "wb") as f:
        f.write(content)

    # Detect period
    dates = [_parse_date(r.txn_date) for r in rows if r.txn_date]
    dates = [d for d in dates if d]
    period_start = min(dates) if dates else None
    period_end = max(dates) if dates else None

    # Create statement record
    r = await db.execute(
        text("""
            INSERT INTO bank_statements (account_id, filename, file_url, period_start, period_end, status, row_count)
            VALUES (:aid, :fname, :furl, :pstart, :pend, 'done', :rcount)
        """),
        {"aid": account_id, "fname": filename, "furl": f"/uploads/bank_statements/{safe_name}",
         "pstart": period_start, "pend": period_end, "rcount": len(rows)},
    )
    await db.commit()
    stmt_id = r.lastrowid

    # Insert transactions
    for row in rows:
        await db.execute(
            text("""
                INSERT INTO bank_transactions
                (account_id, statement_id, txn_date, description, party_name, amount, type, note)
                VALUES (:aid, :sid, :d, :desc, :party, :amt, :type, :note)
            """),
            {"aid": account_id, "sid": stmt_id, "d": row.txn_date,
             "desc": row.description, "party": row.party_name,
             "amt": row.amount, "type": row.type, "note": row.note},
        )
    await db.commit()

    return {"ok": True, "statement_id": stmt_id, "imported": len(rows)}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_cat_groups(ids_str: Optional[str], names_str: Optional[str], colors_str: Optional[str]) -> List[CategoryRef]:
    if not ids_str:
        return []
    ids = [int(x) for x in ids_str.split(",") if x]
    names = names_str.split("|||") if names_str else []
    colors = colors_str.split(",") if colors_str else []
    return [
        CategoryRef(id=ids[i], name=names[i] if i < len(names) else "", color=colors[i] if i < len(colors) else "#6366f1")
        for i in range(len(ids))
    ]


def _parse_inv_groups(ids_str: Optional[str], numbers_str: Optional[str], clients_str: Optional[str]) -> List[InvoiceRef]:
    if not ids_str:
        return []
    ids = [int(x) for x in ids_str.split(",") if x]
    numbers = numbers_str.split("|||") if numbers_str else []
    clients = clients_str.split("|||") if clients_str else []
    return [
        InvoiceRef(id=ids[i], invoice_number=numbers[i] if i < len(numbers) else "", client_name=clients[i] if i < len(clients) else None)
        for i in range(len(ids))
    ]


def _txn_out_from_row(row: tuple) -> TransactionOut:
    cats = _parse_cat_groups(row[8], row[9], row[10])
    invs = _parse_inv_groups(row[11], row[12], row[13])
    first_cat = cats[0] if cats else None
    first_inv = invs[0] if invs else None
    return TransactionOut(
        id=row[0], account_id=row[1], statement_id=row[2],
        txn_date=str(row[3]), description=row[4], party_name=row[5],
        amount=float(row[6]), type=row[7],
        categories=cats,
        category_id=first_cat.id if first_cat else None,
        category_name=first_cat.name if first_cat else None,
        category_color=first_cat.color if first_cat else None,
        invoices=invs,
        invoice_id=first_inv.id if first_inv else None,
        invoice_number=first_inv.invoice_number if first_inv else None,
        bill_id=row[14], bill_number=row[15],
        note=row[16], receipt_url=row[17],
    )


_TXN_SELECT = """
    SELECT bt.id, bt.account_id, bt.statement_id, bt.txn_date, bt.description,
           bt.party_name, bt.amount, bt.type,
           GROUP_CONCAT(DISTINCT btc.category_id ORDER BY tc.name SEPARATOR ',')      AS cat_ids,
           GROUP_CONCAT(DISTINCT tc.name         ORDER BY tc.name SEPARATOR '|||')    AS cat_names,
           GROUP_CONCAT(DISTINCT tc.color        ORDER BY tc.name SEPARATOR ',')      AS cat_colors,
           GROUP_CONCAT(DISTINCT bti.invoice_id  ORDER BY bti.invoice_id SEPARATOR ',')         AS inv_ids,
           GROUP_CONCAT(DISTINCT inv.invoice_number ORDER BY bti.invoice_id SEPARATOR '|||')    AS inv_numbers,
           GROUP_CONCAT(DISTINCT COALESCE(ic.company_name,'') ORDER BY bti.invoice_id SEPARATOR '|||') AS inv_clients,
           bt.bill_id, b.bill_number,
           bt.note, bt.receipt_url
    FROM bank_transactions bt
    LEFT JOIN bank_transaction_categories btc ON btc.transaction_id = bt.id
    LEFT JOIN transaction_categories tc ON tc.id = btc.category_id
    LEFT JOIN bank_transaction_invoices bti ON bti.transaction_id = bt.id
    LEFT JOIN invoices inv ON inv.id = bti.invoice_id
    LEFT JOIN clients ic ON ic.id = inv.client_id
    LEFT JOIN bills b ON b.id = bt.bill_id
"""

_TXN_GROUP = """
    GROUP BY bt.id, bt.account_id, bt.statement_id, bt.txn_date, bt.description,
             bt.party_name, bt.amount, bt.type,
             bt.bill_id, b.bill_number,
             bt.note, bt.receipt_url
"""


async def _upsert_categories(db: AsyncSession, txn_id: int, category_ids: List[int]) -> None:
    await db.execute(
        text("DELETE FROM bank_transaction_categories WHERE transaction_id = :tid"),
        {"tid": txn_id},
    )
    for cat_id in category_ids:
        await db.execute(
            text("INSERT IGNORE INTO bank_transaction_categories (transaction_id, category_id) VALUES (:tid, :cid)"),
            {"tid": txn_id, "cid": cat_id},
        )
    first_cat = category_ids[0] if category_ids else None
    await db.execute(
        text("UPDATE bank_transactions SET category_id = :cid WHERE id = :tid"),
        {"cid": first_cat, "tid": txn_id},
    )


async def _upsert_invoices(db: AsyncSession, txn_id: int, invoice_ids: List[int], txn_type: str) -> None:
    # Get currently linked invoice ids to detect newly added ones
    r = await db.execute(
        text("SELECT invoice_id FROM bank_transaction_invoices WHERE transaction_id = :tid"),
        {"tid": txn_id},
    )
    old_ids = {row[0] for row in r.fetchall()}

    await db.execute(
        text("DELETE FROM bank_transaction_invoices WHERE transaction_id = :tid"),
        {"tid": txn_id},
    )
    for inv_id in invoice_ids:
        await db.execute(
            text("INSERT IGNORE INTO bank_transaction_invoices (transaction_id, invoice_id) VALUES (:tid, :iid)"),
            {"tid": txn_id, "iid": inv_id},
        )
    # Keep legacy column in sync (first invoice or NULL)
    first_inv = invoice_ids[0] if invoice_ids else None
    await db.execute(
        text("UPDATE bank_transactions SET invoice_id = :iid WHERE id = :tid"),
        {"iid": first_inv, "tid": txn_id},
    )
    # Auto-mark newly linked invoices as paid for credit transactions
    if txn_type == "credit":
        for inv_id in invoice_ids:
            if inv_id not in old_ids:
                await db.execute(
                    text("""
                        UPDATE invoices
                        SET status = 'paid', amount_paid = total, balance_due = 0, paid_at = NOW()
                        WHERE id = :id
                    """),
                    {"id": inv_id},
                )


# ── Transaction endpoints ─────────────────────────────────────────────────────

@router.get("/accounts/{account_id}/transactions", response_model=List[TransactionOut])
async def list_transactions(
    account_id: int,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    await _get_account(db, account_id, tid)

    where = ["bt.account_id = :aid"]
    params: dict = {"aid": account_id}

    if date_from:
        where.append("bt.txn_date >= :date_from")
        params["date_from"] = date_from
    if date_to:
        where.append("bt.txn_date <= :date_to")
        params["date_to"] = date_to
    if type:
        where.append("bt.type = :type")
        params["type"] = type
    if category_id:
        where.append("EXISTS (SELECT 1 FROM bank_transaction_categories x WHERE x.transaction_id = bt.id AND x.category_id = :cat_id)")
        params["cat_id"] = category_id
    if search:
        where.append("(bt.description LIKE :s OR bt.party_name LIKE :s)")
        params["s"] = f"%{search}%"

    sql = f"{_TXN_SELECT} WHERE {' AND '.join(where)} {_TXN_GROUP} ORDER BY bt.txn_date DESC, bt.id DESC"
    r = await db.execute(text(sql), params)
    return [_txn_out_from_row(row) for row in r.fetchall()]


@router.post("/accounts/{account_id}/transactions", response_model=TransactionOut)
async def create_transaction(
    account_id: int,
    payload: TransactionCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    await _get_account(db, account_id, tid)
    first_cat = payload.category_ids[0] if payload.category_ids else None
    r = await db.execute(
        text("""
            INSERT INTO bank_transactions
            (account_id, txn_date, description, party_name, amount, type, category_id, note)
            VALUES (:aid, :d, :desc, :party, :amt, :type, :cat, :note)
        """),
        {"aid": account_id, "d": payload.txn_date, "desc": payload.description,
         "party": payload.party_name, "amt": payload.amount, "type": payload.type,
         "cat": first_cat, "note": payload.note},
    )
    txn_id = r.lastrowid
    if payload.category_ids:
        await _upsert_categories(db, txn_id, payload.category_ids)
    if payload.invoice_ids:
        await _upsert_invoices(db, txn_id, payload.invoice_ids, payload.type)
    await db.commit()

    r2 = await db.execute(
        text(f"{_TXN_SELECT} WHERE bt.id = :id {_TXN_GROUP}"),
        {"id": txn_id},
    )
    return _txn_out_from_row(r2.fetchone())


@router.put("/transactions/{txn_id}", response_model=TransactionOut)
async def update_transaction(
    txn_id: int,
    payload: TransactionUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    # Verify ownership via account
    r = await db.execute(
        text("""
            SELECT bt.id, bt.account_id, bt.statement_id, bt.txn_date, bt.description,
                   bt.party_name, bt.amount, bt.type, bt.category_id,
                   bt.invoice_id, bt.bill_id, bt.note
            FROM bank_transactions bt
            JOIN bank_accounts ba ON ba.id = bt.account_id
            WHERE bt.id = :id AND ba.tenant_id = :tid
        """),
        {"id": txn_id, "tid": tid},
    )
    row = r.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Transaction not found")

    old = {
        "txn_date": row[3], "description": row[4], "party_name": row[5],
        "amount": row[6], "type": row[7], "category_id": row[8],
        "invoice_id": row[9], "bill_id": row[10], "note": row[11],
    }
    new_bill_id = payload.bill_id if payload.bill_id is not None else old["bill_id"]
    new_amount = payload.amount if payload.amount is not None else float(old["amount"])
    new_type = payload.type if payload.type is not None else old["type"]

    if payload.category_ids is not None:
        new_cat_id = payload.category_ids[0] if payload.category_ids else None
    else:
        new_cat_id = old["category_id"]

    await db.execute(
        text("""
            UPDATE bank_transactions
            SET txn_date    = :d,
                description = :desc,
                party_name  = :party,
                amount      = :amt,
                type        = :type,
                category_id = :cat,
                bill_id     = :bill,
                note        = :note
            WHERE id = :id
        """),
        {
            "d": payload.txn_date or old["txn_date"],
            "desc": payload.description or old["description"],
            "party": payload.party_name if payload.party_name is not None else old["party_name"],
            "amt": new_amount,
            "type": new_type,
            "cat": new_cat_id,
            "bill": new_bill_id,
            "note": payload.note if payload.note is not None else old["note"],
            "id": txn_id,
        },
    )

    if payload.category_ids is not None:
        await _upsert_categories(db, txn_id, payload.category_ids)

    if payload.invoice_ids is not None:
        await _upsert_invoices(db, txn_id, payload.invoice_ids, new_type)

    # Auto-mark bill as paid when linked
    if new_bill_id and new_bill_id != old["bill_id"] and new_type == "debit":
        await db.execute(
            text("""
                UPDATE bills
                SET status = 'paid', paid_at = NOW()
                WHERE id = :bill_id
            """),
            {"bill_id": new_bill_id},
        )

    await db.commit()

    r2 = await db.execute(
        text(f"{_TXN_SELECT} WHERE bt.id = :id {_TXN_GROUP}"),
        {"id": txn_id},
    )
    return _txn_out_from_row(r2.fetchone())


@router.delete("/transactions/{txn_id}")
async def delete_transaction(
    txn_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    r = await db.execute(
        text("""
            SELECT bt.id FROM bank_transactions bt
            JOIN bank_accounts ba ON ba.id = bt.account_id
            WHERE bt.id = :id AND ba.tenant_id = :tid
        """),
        {"id": txn_id, "tid": tid},
    )
    if not r.fetchone():
        raise HTTPException(status_code=404, detail="Transaction not found")
    await db.execute(text("DELETE FROM bank_transactions WHERE id = :id"), {"id": txn_id})
    await db.commit()
    return {"ok": True}


# ── Receipt upload ────────────────────────────────────────────────────────────

@router.post("/transactions/{txn_id}/receipt")
async def upload_receipt(
    txn_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload a receipt/invoice document for a debit transaction."""
    tid = _tenant_id(current_user)
    r = await db.execute(
        text("""
            SELECT bt.id FROM bank_transactions bt
            JOIN bank_accounts ba ON ba.id = bt.account_id
            WHERE bt.id = :id AND ba.tenant_id = :tid
        """),
        {"id": txn_id, "tid": tid},
    )
    if not r.fetchone():
        raise HTTPException(status_code=404, detail="Transaction not found")

    content = await file.read()
    filename = file.filename or "receipt"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "bin"
    if ext not in {"pdf", "jpg", "jpeg", "png", "gif", "webp"}:
        raise HTTPException(status_code=400, detail="Only PDF or image files are supported")

    upload_dir = os.environ.get("UPLOAD_DIR", "uploads")
    safe_name = f"{txn_id}_{secrets.token_hex(6)}.{ext}"
    file_path = f"{upload_dir}/bank_receipts/{safe_name}"
    with open(file_path, "wb") as f:
        f.write(content)

    receipt_url = f"/uploads/bank_receipts/{safe_name}"
    await db.execute(
        text("UPDATE bank_transactions SET receipt_url = :url WHERE id = :id"),
        {"url": receipt_url, "id": txn_id},
    )
    await db.commit()
    return {"receipt_url": receipt_url}


# ── Cash-flow summary ─────────────────────────────────────────────────────────

@router.get("/accounts/{account_id}/summary")
async def get_summary(
    account_id: int,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = _tenant_id(current_user)
    row = await _get_account(db, account_id, tid)

    # Opening balance
    opening_r = await db.execute(
        text("SELECT opening_balance, currency FROM bank_accounts WHERE id = :id"),
        {"id": account_id},
    )
    acc_row = opening_r.fetchone()
    opening_balance = float(acc_row[0])
    currency = acc_row[1]

    where = ["account_id = :aid"]
    params: dict = {"aid": account_id}
    if date_from:
        where.append("txn_date >= :df")
        params["df"] = date_from
    if date_to:
        where.append("txn_date <= :dt")
        params["dt"] = date_to
    if category_id:
        where.append("id IN (SELECT transaction_id FROM bank_transaction_categories WHERE category_id = :cat_id)")
        params["cat_id"] = category_id

    # Overall totals
    totals_r = await db.execute(
        text(f"""
            SELECT
                COALESCE(SUM(CASE WHEN type='credit' THEN amount ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN type='debit'  THEN amount ELSE 0 END), 0)
            FROM bank_transactions WHERE {' AND '.join(where)}
        """),
        params,
    )
    tot = totals_r.fetchone()
    total_credit = float(tot[0])
    total_debit = float(tot[1])
    net = total_credit - total_debit

    # Full running balance uses all transactions (not filtered by date)
    _, _, current_balance = await _account_balances(db, account_id, opening_balance)

    # Monthly breakdown
    monthly_r = await db.execute(
        text(f"""
            SELECT
                DATE_FORMAT(txn_date, '%Y-%m') AS month,
                COALESCE(SUM(CASE WHEN type='credit' THEN amount ELSE 0 END), 0) AS credit,
                COALESCE(SUM(CASE WHEN type='debit'  THEN amount ELSE 0 END), 0) AS debit
            FROM bank_transactions
            WHERE {' AND '.join(where)}
            GROUP BY month
            ORDER BY month
        """),
        params,
    )
    monthly = [
        {"month": r[0], "credit": float(r[1]), "debit": float(r[2]), "net": float(r[1]) - float(r[2])}
        for r in monthly_r.fetchall()
    ]

    # Category breakdown (uses junction table; maps credit→income, debit→expense)
    cat_r = await db.execute(
        text(f"""
            SELECT tc.name, tc.color,
                   CASE WHEN bt.type = 'credit' THEN 'income' ELSE 'expense' END AS cat_type,
                   COALESCE(SUM(bt.amount), 0) AS total
            FROM bank_transactions bt
            LEFT JOIN bank_transaction_categories btc ON btc.transaction_id = bt.id
            LEFT JOIN transaction_categories tc ON tc.id = btc.category_id
            WHERE {' AND '.join(where)}
            GROUP BY btc.category_id, tc.name, tc.color, bt.type
            ORDER BY total DESC
        """),
        params,
    )
    categories = [
        {"name": r[0] or "Uncategorised", "color": r[1] or "#94a3b8", "type": r[2], "total": float(r[3])}
        for r in cat_r.fetchall()
    ]

    return {
        "currency": currency,
        "opening_balance": opening_balance,
        "current_balance": round(current_balance, 2),
        "total_credit": round(total_credit, 2),
        "total_debit": round(total_debit, 2),
        "net": round(net, 2),
        "monthly": monthly,
        "categories": categories,
    }


# ── Invoice/Bill lookup helpers for reconciliation ───────────────────────────

@router.get("/invoices/unpaid")
async def list_unpaid_invoices(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    txn_id: Optional[int] = None,
):
    tid = get_effective_tenant_id(current_user)
    where = "i.status != 'cancelled' AND (i.is_deleted = 0 OR i.is_deleted IS NULL)"
    params: dict = {}
    if tid is not None:
        where += " AND i.tenant_id = :tid"
        params["tid"] = tid
    if txn_id is not None:
        # Exclude invoices already linked to a different transaction
        where += """
            AND (
                NOT EXISTS (SELECT 1 FROM bank_transaction_invoices bti WHERE bti.invoice_id = i.id)
                OR EXISTS (SELECT 1 FROM bank_transaction_invoices bti WHERE bti.invoice_id = i.id AND bti.transaction_id = :txn_id)
            )
        """
        params["txn_id"] = txn_id
    r = await db.execute(
        text(f"""
            SELECT i.id, i.invoice_number, i.total, i.balance_due, i.status, c.company_name AS client_name
            FROM invoices i
            LEFT JOIN clients c ON c.id = i.client_id
            WHERE {where}
            ORDER BY i.issue_date DESC
            LIMIT 500
        """),
        params,
    )
    return [
        {"id": row[0], "invoice_number": row[1], "total": float(row[2]),
         "balance_due": float(row[3]), "client_name": row[5] or "—", "status": row[4]}
        for row in r.fetchall()
    ]


@router.get("/bills/unpaid")
async def list_unpaid_bills(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = get_effective_tenant_id(current_user)
    where = "is_deleted = 0"
    params: dict = {}
    if tid is not None:
        where += " AND tenant_id = :tid"
        params["tid"] = tid
    r = await db.execute(
        text(f"""
            SELECT id, bill_number, amount, vendor_name, status
            FROM bills
            WHERE {where}
            ORDER BY issue_date DESC
            LIMIT 200
        """),
        params,
    )
    return [
        {"id": row[0], "bill_number": row[1], "amount": float(row[2]),
         "vendor_name": row[3] or "—", "status": row[4]}
        for row in r.fetchall()
    ]


@router.get("/party/search")
async def search_parties(
    q: str = Query(""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Search clients, vendors, and staff by name. Returns up to 10 results per source."""
    tid = get_effective_tenant_id(current_user)
    term = f"%{q.strip()}%"
    tenant_filter = "AND tenant_id = :tid" if tid is not None else ""
    params: dict = {"term": term}
    if tid is not None:
        params["tid"] = tid

    sql = text(f"""
        (SELECT 'client' AS source, company_name AS name FROM clients
         WHERE company_name LIKE :term {tenant_filter} LIMIT 10)
        UNION ALL
        (SELECT 'vendor' AS source, name AS name FROM vendors
         WHERE name LIKE :term {tenant_filter} LIMIT 10)
        UNION ALL
        (SELECT 'staff' AS source, full_name AS name FROM hr_employees
         WHERE full_name LIKE :term {tenant_filter} LIMIT 10)
    """)
    r = await db.execute(sql, params)
    return [{"source": row[0], "name": row[1]} for row in r.fetchall()]


# ── Cashflow report helpers ───────────────────────────────────────────────────

async def _cashflow_data(db: AsyncSession, date_from: str, date_to: str, tid: Optional[int]):
    """Fetch all data needed for cashflow reports (PDF + Excel share this)."""
    tenant_filter = "AND ba.tenant_id = :tid" if tid is not None else ""
    params: dict = {"d_from": date_from, "d_to": date_to}
    if tid is not None:
        params["tid"] = tid

    # Opening balance = sum of account opening balances + all transactions before date_from
    r = await db.execute(
        text(f"""
            SELECT
                COALESCE(SUM(ba.opening_balance), 0) +
                COALESCE(SUM(CASE WHEN bt.txn_date < :d_from AND bt.type = 'credit' THEN bt.amount ELSE 0 END), 0) -
                COALESCE(SUM(CASE WHEN bt.txn_date < :d_from AND bt.type = 'debit'  THEN bt.amount ELSE 0 END), 0)
            FROM bank_accounts ba
            LEFT JOIN bank_transactions bt ON bt.account_id = ba.id
            WHERE 1=1 {tenant_filter}
        """),
        params,
    )
    opening_balance = float(r.scalar() or 0)

    # All transactions in range
    r = await db.execute(
        text(f"""
            SELECT bt.txn_date, bt.description, bt.party_name, bt.amount, bt.type,
                   ba.name AS account_name,
                   GROUP_CONCAT(DISTINCT tc.name ORDER BY tc.name SEPARATOR ', ') AS categories
            FROM bank_transactions bt
            JOIN bank_accounts ba ON ba.id = bt.account_id
            LEFT JOIN bank_transaction_categories btc ON btc.transaction_id = bt.id
            LEFT JOIN transaction_categories tc ON tc.id = btc.category_id
            WHERE bt.txn_date BETWEEN :d_from AND :d_to {tenant_filter}
            GROUP BY bt.id, bt.txn_date, bt.description, bt.party_name, bt.amount, bt.type, ba.name
            ORDER BY bt.txn_date ASC, bt.id ASC
        """),
        params,
    )
    rows = r.fetchall()

    # Build transaction list with running balance
    transactions = []
    running = opening_balance
    for row in rows:
        if row[4] == "credit":
            running += float(row[2] if isinstance(row[2], float) else row[3])
        else:
            running -= float(row[2] if isinstance(row[2], float) else row[3])
        # row: txn_date, description, party_name, amount, type, account_name, categories
        amt = float(row[3])
        running_prev = running
        if row[4] == "credit":
            running = running_prev  # already added above
        transactions.append({
            "txn_date": str(row[0]),
            "description": row[1] or "",
            "party_name": row[2] or "",
            "amount": float(row[3]),
            "type": row[4],
            "account_name": row[5] or "",
            "categories": row[6] or "",
            "running_balance": running,
        })

    # Recalculate running balance correctly
    running = opening_balance
    for t in transactions:
        if t["type"] == "credit":
            running += t["amount"]
        else:
            running -= t["amount"]
        t["running_balance"] = running

    total_credit = sum(t["amount"] for t in transactions if t["type"] == "credit")
    total_debit = sum(t["amount"] for t in transactions if t["type"] == "debit")
    closing_balance = opening_balance + total_credit - total_debit

    # Monthly summary
    from collections import defaultdict
    monthly_map: dict = defaultdict(lambda: {"credit": 0.0, "debit": 0.0})
    for t in transactions:
        ym = t["txn_date"][:7]  # YYYY-MM
        if t["type"] == "credit":
            monthly_map[ym]["credit"] += t["amount"]
        else:
            monthly_map[ym]["debit"] += t["amount"]

    monthly = []
    bal = opening_balance
    MONTHS = ["", "January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    for ym in sorted(monthly_map.keys()):
        year, mon = int(ym[:4]), int(ym[5:])
        cr = monthly_map[ym]["credit"]
        dr = monthly_map[ym]["debit"]
        net = cr - dr
        bal += net
        monthly.append({
            "month_label": f"{MONTHS[mon]} {year}",
            "credit": cr,
            "debit": dr,
            "net": net,
            "balance": bal,
        })

    return {
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "total_credit": total_credit,
        "total_debit": total_debit,
        "transactions": transactions,
        "monthly": monthly,
    }


@router.get("/accounts/{account_id}/transactions/report")
async def account_transactions_pdf(
    account_id: int,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from fastapi.responses import Response
    from jinja2 import Environment, FileSystemLoader
    from weasyprint import HTML
    from pathlib import Path

    tid = _tenant_id(current_user)
    await _get_account(db, account_id, tid)  # access check only
    acc_row = await db.execute(
        text("SELECT name, bank_name, account_number, currency FROM bank_accounts WHERE id = :id"),
        {"id": account_id},
    )
    acc = acc_row.fetchone()

    where = ["bt.account_id = :aid"]
    params: dict = {"aid": account_id}
    if date_from:
        where.append("bt.txn_date >= :date_from")
        params["date_from"] = date_from
    if date_to:
        where.append("bt.txn_date <= :date_to")
        params["date_to"] = date_to
    if type:
        where.append("bt.type = :type")
        params["type"] = type
    if category_id:
        where.append("EXISTS (SELECT 1 FROM bank_transaction_categories x WHERE x.transaction_id = bt.id AND x.category_id = :cat_id)")
        params["cat_id"] = category_id
    if search:
        where.append("(bt.description LIKE :s OR bt.party_name LIKE :s)")
        params["s"] = f"%{search}%"

    sql = f"{_TXN_SELECT} WHERE {' AND '.join(where)} {_TXN_GROUP} ORDER BY bt.txn_date ASC, bt.id ASC"
    r = await db.execute(text(sql), params)
    txns = [_txn_out_from_row(row) for row in r.fetchall()]

    total_credit = sum(t.amount for t in txns if t.type == "credit")
    total_debit = sum(t.amount for t in txns if t.type == "debit")

    # Company settings
    cs = await db.execute(
        text("SELECT * FROM company_settings WHERE (tenant_id=:tid OR (tenant_id IS NULL AND :tid IS NULL)) LIMIT 1"),
        {"tid": tid}
    )
    company = cs.mappings().first() or {}
    logo_data = company.get("logo_url")
    primary_color = company.get("primary_color") or "#1a1a2e"

    # Build filter label
    filters = []
    if date_from or date_to:
        filters.append(f"{date_from or '—'} to {date_to or '—'}")
    if type:
        filters.append(f"Type: {type.capitalize()}")
    if search:
        filters.append(f'Search: "{search}"')

    context = {
        "company": company,
        "logo_data": logo_data,
        "primary_color": primary_color,
        "account_name": acc[0] if acc else "",
        "account_bank": acc[1] or "" if acc else "",
        "account_number": acc[2] or "" if acc else "",
        "currency": acc[3] if acc else "MYR",
        "filter_label": " | ".join(filters) if filters else "All transactions",
        "generated_at": date.today().strftime("%d %B %Y"),
        "txns": [
            {
                "txn_date": str(t.txn_date) if hasattr(t.txn_date, "strftime") else t.txn_date,
                "description": t.description or "",
                "party_name": t.party_name or "",
                "categories": ", ".join(c.name for c in (t.categories or [])),
                "type": t.type,
                "amount": float(t.amount),
            }
            for t in txns
        ],
        "total_credit": total_credit,
        "total_debit": total_debit,
        "net": total_credit - total_debit,
        "count": len(txns),
    }

    templates_dir = Path(__file__).parent.parent / "templates" / "pdf"
    jinja_env = Environment(loader=FileSystemLoader(str(templates_dir)))
    template = jinja_env.get_template("account_report.html")
    html_content = template.render(**context)

    import asyncio
    loop = asyncio.get_event_loop()
    pdf_bytes = await loop.run_in_executor(None, lambda: HTML(string=html_content).write_pdf())

    fname = f"report_{acc[0] if acc else account_id}_{date_from or 'all'}_{date_to or 'all'}.pdf".replace(" ", "_")
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/accounts/{account_id}/transactions/excel")
async def account_transactions_excel(
    account_id: int,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    tid = _tenant_id(current_user)
    await _get_account(db, account_id, tid)  # access check only
    acc_row2 = await db.execute(
        text("SELECT name, bank_name, account_number, currency FROM bank_accounts WHERE id = :id"),
        {"id": account_id},
    )
    acc2 = acc_row2.fetchone()

    where = ["bt.account_id = :aid"]
    params: dict = {"aid": account_id}
    if date_from:
        where.append("bt.txn_date >= :date_from")
        params["date_from"] = date_from
    if date_to:
        where.append("bt.txn_date <= :date_to")
        params["date_to"] = date_to
    if type:
        where.append("bt.type = :type")
        params["type"] = type
    if category_id:
        where.append("EXISTS (SELECT 1 FROM bank_transaction_categories x WHERE x.transaction_id = bt.id AND x.category_id = :cat_id)")
        params["cat_id"] = category_id
    if search:
        where.append("(bt.description LIKE :s OR bt.party_name LIKE :s)")
        params["s"] = f"%{search}%"

    sql = f"{_TXN_SELECT} WHERE {' AND '.join(where)} {_TXN_GROUP} ORDER BY bt.txn_date ASC, bt.id ASC"
    r = await db.execute(text(sql), params)
    txns = [_txn_out_from_row(row) for row in r.fetchall()]

    navy = "1A1A2E"
    green = "065F46"
    red_c = "991B1B"
    num_fmt = '#,##0.00'
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor=navy)
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    title_font = Font(name="Arial", bold=True, size=13, color=navy)
    sub_font = Font(name="Arial", size=9, color="555555")

    ws["A1"] = f"Transaction Report — {acc2[0] if acc2 else account_id}"
    ws["A1"].font = title_font
    ws["A2"] = f"Period: {date_from or 'all'} to {date_to or 'all'}"
    ws["A2"].font = sub_font

    headers = ["Date", "Description", "Party", "Category", "Remarks", "Money In", "Money Out"]
    h_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(h_row, col)
        cell.value = h
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr

    total_credit = 0.0
    total_debit = 0.0
    for i, t in enumerate(txns):
        r = h_row + 1 + i
        ws.cell(r, 1).value = str(t.txn_date) if hasattr(t.txn_date, "strftime") else t.txn_date
        ws.cell(r, 2).value = t.description or ""
        ws.cell(r, 3).value = t.party_name or ""
        ws.cell(r, 4).value = ", ".join(c.name for c in (t.categories or []))
        ws.cell(r, 5).value = t.note or ""
        is_credit = t.type == "credit"
        ws.cell(r, 6).value = float(t.amount) if is_credit else None
        ws.cell(r, 7).value = float(t.amount) if not is_credit else None
        if is_credit:
            total_credit += float(t.amount)
            ws.cell(r, 6).font = Font(name="Arial", size=9, color=green)
        else:
            total_debit += float(t.amount)
            ws.cell(r, 7).font = Font(name="Arial", size=9, color=red_c)
        for col in range(1, 8):
            ws.cell(r, col).font = ws.cell(r, col).font or Font(name="Arial", size=9)
            ws.cell(r, col).border = bdr
            fill = PatternFill("solid", fgColor="F8F8F8" if i % 2 else "FFFFFF")
            ws.cell(r, col).fill = fill
        for col in [6, 7]:
            ws.cell(r, col).number_format = num_fmt

    # Total row
    tr = h_row + 1 + len(txns)
    ws.cell(tr, 1).value = "TOTAL"
    ws.cell(tr, 6).value = total_credit
    ws.cell(tr, 7).value = total_debit
    for col in range(1, 8):
        ws.cell(tr, col).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        ws.cell(tr, col).fill = PatternFill("solid", fgColor=navy)
        ws.cell(tr, col).border = bdr
    for col in [6, 7]:
        ws.cell(tr, col).number_format = num_fmt

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"report_{acc2[0] if acc2 else account_id}_{date_from or 'all'}_{date_to or 'all'}.xlsx".replace(" ", "_")
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/cashflow/pdf")
async def cashflow_pdf(
    date_from: str = Query(...),
    date_to: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from io import BytesIO
    import base64
    from jinja2 import Environment, FileSystemLoader
    from weasyprint import HTML as WP_HTML

    tid = get_effective_tenant_id(current_user)
    data = await _cashflow_data(db, date_from, date_to, tid)

    # Company info + logo
    r = await db.execute(text("SELECT name, address, phone, email, logo_url FROM company_settings WHERE tenant_id = :tid LIMIT 1"), {"tid": tid})
    cr = r.fetchone()
    company = {"name": cr[0], "address": cr[1], "phone": cr[2], "email": cr[3]} if cr else {}
    logo_data = None
    if cr and cr[4]:
        logo_path = os.path.join("uploads", cr[4].lstrip("/uploads/").lstrip("/"))
        if os.path.exists(logo_path):
            with open(logo_path, "rb") as f:
                ext = logo_path.rsplit(".", 1)[-1].lower()
                mime = "image/png" if ext == "png" else "image/jpeg"
                logo_data = f"data:{mime};base64,{base64.b64encode(f.read()).decode()}"

    tmpl_dir = os.path.join(os.path.dirname(__file__), "..", "templates", "pdf")
    env = Environment(loader=FileSystemLoader(tmpl_dir))
    tmpl = env.get_template("cashflow.html")
    html = tmpl.render(
        company=company,
        logo_data=logo_data,
        date_from=date_from,
        date_to=date_to,
        generated_at=datetime.now().strftime("%d %b %Y %H:%M"),
        **data,
    )
    from fastapi.responses import StreamingResponse
    pdf_bytes = WP_HTML(string=html).write_pdf()
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="cashflow_{date_from}_{date_to}.pdf"'},
    )


@router.get("/cashflow/excel")
async def cashflow_excel(
    date_from: str = Query(...),
    date_to: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    tid = get_effective_tenant_id(current_user)
    data = await _cashflow_data(db, date_from, date_to, tid)

    wb = openpyxl.Workbook()

    # ── Styles ──────────────────────────────────────────────────────────────
    navy = "1A1A2E"
    green = "065F46"
    red = "991B1B"
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor=navy)
    hdr_align = Alignment(horizontal="center", vertical="center")
    title_font = Font(name="Arial", bold=True, size=13, color=navy)
    sub_font = Font(name="Arial", size=9, color="444444")
    num_fmt = '#,##0.00'
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = border

    def style_data_row(ws, row, cols, even=False):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", fgColor="F8F8F8" if even else "FFFFFF")
            cell.border = border

    # ── Sheet 1: Monthly Summary ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Monthly Summary"

    ws1["A1"] = "Cash Flow Statement"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Period: {date_from} to {date_to}"
    ws1["A2"].font = sub_font
    ws1["A3"] = f"Opening Balance: MYR {data['opening_balance']:,.2f}"
    ws1["A3"].font = Font(name="Arial", size=10, bold=True, color=navy)

    # Balance summary row
    ws1["A5"] = "Opening Balance"
    ws1["B5"] = data["opening_balance"]
    ws1["A6"] = "Total Cash In"
    ws1["B6"] = data["total_credit"]
    ws1["A7"] = "Total Cash Out"
    ws1["B7"] = data["total_debit"]
    ws1["A8"] = "Closing Balance"
    ws1["B8"] = data["closing_balance"]
    for r in range(5, 9):
        ws1.cell(r, 1).font = Font(name="Arial", bold=True, size=9)
        ws1.cell(r, 2).number_format = num_fmt
        ws1.cell(r, 2).font = Font(name="Arial", size=9)

    # Monthly table
    headers = ["Month", "Cash In (MYR)", "Cash Out (MYR)", "Net (MYR)", "Closing Balance (MYR)"]
    start_row = 10
    for col, h in enumerate(headers, 1):
        ws1.cell(start_row, col).value = h
    style_header_row(ws1, start_row, len(headers))

    for i, row in enumerate(data["monthly"]):
        r = start_row + 1 + i
        ws1.cell(r, 1).value = row["month_label"]
        ws1.cell(r, 2).value = row["credit"]
        ws1.cell(r, 3).value = row["debit"]
        ws1.cell(r, 4).value = row["net"]
        ws1.cell(r, 5).value = row["balance"]
        style_data_row(ws1, r, len(headers), i % 2 == 1)
        for col in [2, 3, 4, 5]:
            ws1.cell(r, col).number_format = num_fmt
        ws1.cell(r, 4).font = Font(name="Arial", size=9,
                                   color=green if row["net"] >= 0 else red, bold=True)

    # Total row
    tr = start_row + 1 + len(data["monthly"])
    ws1.cell(tr, 1).value = "TOTAL"
    ws1.cell(tr, 2).value = data["total_credit"]
    ws1.cell(tr, 3).value = data["total_debit"]
    ws1.cell(tr, 4).value = data["total_credit"] - data["total_debit"]
    ws1.cell(tr, 5).value = data["closing_balance"]
    style_header_row(ws1, tr, len(headers))
    for col in [2, 3, 4, 5]:
        ws1.cell(tr, col).number_format = num_fmt

    ws1.column_dimensions["A"].width = 20
    for col in ["B", "C", "D", "E"]:
        ws1.column_dimensions[col].width = 22

    # ── Sheet 2: Transaction Details ─────────────────────────────────────────
    ws2 = wb.create_sheet("Transaction Details")
    headers2 = ["Date", "Account", "Description", "Party", "Category", "Cash In (MYR)", "Cash Out (MYR)", "Balance (MYR)"]
    ws2["A1"] = "Transaction Details"
    ws2["A1"].font = title_font
    ws2["A2"] = f"Period: {date_from} to {date_to}"
    ws2["A2"].font = sub_font

    h_row = 4
    for col, h in enumerate(headers2, 1):
        ws2.cell(h_row, col).value = h
    style_header_row(ws2, h_row, len(headers2))

    for i, t in enumerate(data["transactions"]):
        r = h_row + 1 + i
        ws2.cell(r, 1).value = t["txn_date"]
        ws2.cell(r, 2).value = t["account_name"]
        ws2.cell(r, 3).value = t["description"]
        ws2.cell(r, 4).value = t["party_name"]
        ws2.cell(r, 5).value = t["categories"]
        ws2.cell(r, 6).value = t["amount"] if t["type"] == "credit" else None
        ws2.cell(r, 7).value = t["amount"] if t["type"] == "debit" else None
        ws2.cell(r, 8).value = t["running_balance"]
        style_data_row(ws2, r, len(headers2), i % 2 == 1)
        for col in [6, 7, 8]:
            ws2.cell(r, col).number_format = num_fmt
        if t["type"] == "credit":
            ws2.cell(r, 6).font = Font(name="Arial", size=9, color=green)
        else:
            ws2.cell(r, 7).font = Font(name="Arial", size=9, color=red)

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 18
    for col in ["F", "G", "H"]:
        ws2.column_dimensions[col].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="cashflow_{date_from}_{date_to}.xlsx"'},
    )
